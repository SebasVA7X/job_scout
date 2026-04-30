"""
notifier/telegram_bot.py
Sends job alerts via Telegram Bot API.

Two notification modes:
- Immediate alert: score >= score_alert_threshold (90+) → individual message per job
- Digest: score >= score_digest_threshold (50+) → Excel file attached once per run
- No new jobs: plain text message so you know the pipeline ran OK
"""
import io
import logging
from datetime import datetime
from typing import List, Dict, Any

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from config.settings import settings

logger = logging.getLogger(__name__)

TELEGRAM_API = f"https://api.telegram.org/bot{settings.telegram_token}"

# Score color bands for Excel
COLOR_HIGH   = "2ECC71"  # green  — 80+
COLOR_MEDIUM = "F39C12"  # orange — 60-79
COLOR_LOW    = "95A5A6"  # grey   — 50-59


def _send_message(text: str) -> bool:
    if not settings.telegram_token or not settings.telegram_chat_id:
        logger.warning("[Telegram] Token or chat_id not configured — skipping.")
        return False
    try:
        resp = requests.post(
            f"{TELEGRAM_API}/sendMessage",
            json={
                "chat_id":                settings.telegram_chat_id,
                "text":                   text,
                "parse_mode":             "HTML",
                "disable_web_page_preview": True,
            },
            timeout=15,
        )
        resp.raise_for_status()
        return True
    except requests.RequestException as e:
        logger.error(f"[Telegram] Failed to send message: {e}")
        return False


def _send_document(file_bytes: bytes, filename: str, caption: str = "") -> bool:
    if not settings.telegram_token or not settings.telegram_chat_id:
        return False
    try:
        resp = requests.post(
            f"{TELEGRAM_API}/sendDocument",
            data={"chat_id": settings.telegram_chat_id, "caption": caption, "parse_mode": "HTML"},
            files={"document": (filename, file_bytes,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
        resp.raise_for_status()
        return True
    except requests.RequestException as e:
        logger.error(f"[Telegram] Failed to send document: {e}")
        return False


# ── Individual alert (90+) ─────────────────────────────────────────────────────
def format_job_alert(job: Dict[str, Any]) -> str:
    remote_tag = "🌍 Remote" if job.get("is_remote") else f"📍 {job.get('location', 'Unknown')}"
    salary = ""
    if job.get("salary_min") or job.get("salary_max"):
        lo = f"${job['salary_min']:,.0f}" if job.get("salary_min") else ""
        hi = f"${job['salary_max']:,.0f}" if job.get("salary_max") else ""
        salary = f"\n💰 <b>Salary:</b> {' – '.join(filter(None, [lo, hi]))} {job.get('salary_currency', '')}"

    return (
        f"🚨 <b>Top Match — Score {job.get('llm_score', '?')}/100</b>\n\n"
        f"<b>{job.get('title_clean') or job.get('title', 'N/A')}</b>\n"
        f"🏢 {job.get('company', 'N/A')}  |  {remote_tag}{salary}\n"
        f"📂 Source: <i>{job.get('source', '')}</i>\n\n"
        f"💬 <i>{job.get('llm_reasoning', '')}</i>\n\n"
        f"🔗 <a href='{job['url']}'>View Job</a>"
    )


def send_immediate_alerts(jobs: List[Dict[str, Any]]) -> List[str]:
    """Send individual alerts for score >= alert_threshold. Returns notified job_ids."""
    notified = []
    for job in jobs:
        import time
        text = format_job_alert(job)
        if _send_message(text):
            notified.append(job["job_id"])
            logger.info(f"[Telegram] Alert sent: {job.get('title_clean') or job.get('title', '?')} @ {job.get('company')} ({job.get('llm_score')})")
        time.sleep(1)  # respect Telegram rate limit (1 msg/sec)
    return notified


# ── Excel digest (50-89) ───────────────────────────────────────────────────────
def _score_fill(score: int) -> PatternFill:
    if score >= 80:
        color = COLOR_HIGH
    elif score >= 60:
        color = COLOR_MEDIUM
    else:
        color = COLOR_LOW
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _build_excel(jobs: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Digest"

    # Header
    headers = ["Score", "Confidence", "Title", "Company", "Location", "Remote", "Source", "Posted", "URL", "LLM Reasoning"]
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Rows — sorted by score descending
    for row, job in enumerate(sorted(jobs, key=lambda j: j.get("llm_score", 0), reverse=True), 2):
        score = job.get("llm_score", 0)
        fill  = _score_fill(score)

        values = [
            score,
            job.get("llm_confidence", "low"),
            job.get("title_clean", ""),
            job.get("company", ""),
            job.get("location", ""),
            "Yes" if job.get("is_remote") else "No",
            job.get("source", ""),
            job.get("date_posted", ""),
            job.get("url", ""),
            job.get("llm_reasoning", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            # Make URL column a hyperlink
            if col == 9 and val:
                cell.hyperlink = val
                cell.font = Font(color="0000FF", underline="single")

    # Column widths
    widths = [7, 10, 45, 25, 25, 8, 15, 12, 50, 60]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_digest(jobs: List[Dict[str, Any]], run_timestamp: str) -> List[str]:
    """
    Build and send Excel digest for jobs in the 50-89 range.
    Returns notified job_ids.
    """
    if not jobs:
        return []

    filename = f"job_digest_{run_timestamp}.xlsx"
    caption  = f"📊 <b>Job Digest</b> — {len(jobs)} vacante(s) | {run_timestamp}"

    try:
        excel_bytes = _build_excel(jobs)
        if _send_document(excel_bytes, filename, caption):
            logger.info(f"[Telegram] Digest sent: {len(jobs)} jobs")
            return [j["job_id"] for j in jobs]
    except Exception as e:
        logger.error(f"[Telegram] Failed to build/send digest: {e}")

    return []


# ── No new jobs notice ─────────────────────────────────────────────────────────
def send_no_new_jobs() -> None:
    _send_message("🤖 <b>Job Scout:</b> No new jobs this run.")


# ── System messages ────────────────────────────────────────────────────────────
def send_system_message(text: str) -> None:
    _send_message(f"🤖 <b>Job Scout:</b> {text}")